import openpyxl

def get_excel_data(path,sheet_name):
    """load workbook"""
    workbook= openpyxl.load_workbook(path)
    """set sheet name from opened excel"""
    sheet= workbook[sheet_name]

    """ store all rows"""
    data= []

    """start from row 2 skip header"""
    for r in range(2,sheet.max_row+1):
        username= sheet.cell(row=r,column=1).value
        password= sheet.cell(row=r,column=2).value

        data.append((username,password)) # append tuple

    return data